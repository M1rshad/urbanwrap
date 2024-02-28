from django.shortcuts import render, HttpResponse, get_object_or_404
from django.http import JsonResponse
# Create your views here.
from django.shortcuts import render,redirect
from django.contrib.auth import login, logout, authenticate
from user_auth.models import User, Coupon
from home.models import Category, Product, Variation , ProductImages
from orders.models import Order, Wallet, OrderProduct, WalletTransaction, Offer
from django.contrib import messages
from django.views.decorators.cache import never_cache
from django.contrib.postgres.search import SearchVector
from user_auth.forms import SignupForm
from .forms import EditUserForm, AddCategoryForm, AddProductForm, AddVariantForm, ProductImageForm, AddCouponForm, AddOfferForm
from orders.forms import OrderUpdateForm, PaymentUpdateForm
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.mail import send_mail
from django.db.models.functions import TruncWeek, TruncMonth, TruncDay, ExtractWeek
from datetime import datetime, timedelta, timezone
from django.utils import timezone as tz
from django.db.models import Sum, F, Q
from .helpers import render_to_pdf
import openpyxl
import openpyxl.styles
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle
import uuid

# Create your views here.
def is_user_admin(user):
    return user.is_authenticated and user.is_superuser


@never_cache
def admin_login(request):
    if request.user.is_authenticated and request.user.is_superuser:
        return redirect(admin_panel)
    if request.POST:
        email = request.POST.get('email')
        password = request.POST.get('password')

        user = authenticate(email=email, password=password)
        if user is not None and user.is_superuser:
            login(request, user)
            return redirect(admin_panel)
        else:
            messages.info(request, 'Please enter the correct email and password for a staff account.')
    return render(request, 'admin_panel/admin_login.html')


@never_cache
@user_passes_test(is_user_admin, login_url='admin_login')
def admin_panel(request):
    all_orders = Order.objects.filter(payment__status='Completed')
    all_order_items = OrderProduct.objects.filter(is_ordered=True)
    all_variants = Variation.objects.all()

    
    
    filter_type = request.GET.get('filter_type', 'all')  

    if filter_type == 'day':
        start_date = datetime.now() - timedelta(days=1)
    elif filter_type == 'week':
        start_date = datetime.now() - timedelta(weeks=1)
    elif filter_type == 'month':
        start_date = datetime.now() - timedelta(weeks=4)  
    elif filter_type == 'year':
        start_date = datetime.now() - timedelta(weeks=52)  
    else:
        start_date = None  
    
    if start_date:
        all_orders = all_orders.filter(
            created_at__gte=start_date,
            is_ordered=True
        ).distinct()
         
        all_order_items = OrderProduct.objects.filter(
            order__in=all_orders,
            is_ordered=True
        )
         
        
    total_revenue = sum(order.order_total for order in all_orders)
    total_sales =sum(item.quantity for item in all_order_items) 
    total_stock = sum(variation.stock for variation in all_variants)

    cod_orders = all_orders.filter(payment_method='cod')
    online_orders = all_orders.filter(payment_method='paypal') 
    wallet_orders = all_orders.filter(payment_method='wallet')

    cod_count = cod_orders.count()
    online_count = online_orders.count()
    wallet_count = wallet_orders.count()

    cod_total = sum(order.order_total for order in cod_orders)
    online_total = sum(order.order_total for order in online_orders)
    wallet_total = sum(order.order_total for order in wallet_orders)
    
    context={
        'total_revenue':total_revenue,
        'total_sales':total_sales,
        'total_stock':total_stock,
        'all_orders':all_orders,
        'all_order_items':all_order_items,
        'cod_count': cod_count,
        'online_count': online_count,
        'wallet_count': wallet_count,
        'cod_total': cod_total,
        'online_total': online_total,
        'wallet_total': wallet_total,
        'filter_type': filter_type,
    }

    return render(request, 'admin_panel/admin_panel.html', context)


def log_out(request):
    logout(request)
    return redirect(admin_login)


@never_cache
@user_passes_test(is_user_admin, login_url='admin_login')
def user_management(request):
    user_obj = User.objects.all().order_by('id')
    context = {'user_obj' : user_obj}
    return render(request, 'admin_panel/user_management.html',context)



@user_passes_test(is_user_admin, login_url='admin_login')
def user_search(request):
    user_obj=None
    if request.POST:
        search_item = request.POST.get('search_input')
        if search_item:
            user_obj = User.objects.order_by('date_joined').filter(Q(email__icontains=search_item) | Q(username__icontains=search_item))
        
        context = {'user_obj':user_obj}
        return render(request, 'admin_panel/user_management.html',context)
    else:
        return redirect(user_management)


@user_passes_test(is_user_admin, login_url='admin_login')
def block_user(request,pk):
    instance = User.objects.get(pk=pk)
    instance.is_block = True
    instance.save()
    return redirect('user_management')


@user_passes_test(is_user_admin, login_url='admin_login')
def unblock_user(request,pk):
    instance = User.objects.get(pk=pk)
    instance.is_block = False
    instance.save()
    return redirect('user_management')


@never_cache
@user_passes_test(is_user_admin, login_url='admin_login')
def category_management(request):
    cat_obj = Category.objects.all().order_by('id')
    context = {'cat_obj' : cat_obj}
    return render(request, 'admin_panel/category_management.html',context)


@user_passes_test(is_user_admin, login_url='admin_login')
def unlist_category(request, pk):
    instance = Category.objects.get(pk=pk)
    instance.is_active = False
    instance.save()
    return redirect('category_management')


@user_passes_test(is_user_admin, login_url='admin_login')
def list_category(request, pk):
    instance = Category.objects.get(pk=pk)
    instance.is_active = True
    instance.save()
    return redirect('category_management')


@user_passes_test(is_user_admin, login_url='admin_login')
def add_category(request):
    form = AddCategoryForm()
    if request.POST:
        form = AddCategoryForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect(category_management)
    context = {'form' : form}
    return render(request, 'admin_panel/add_category.html',context)


@user_passes_test(is_user_admin, login_url='admin_login')
def edit_category(request, pk):
    error_category_name = ''
    error_slug = ''
    error_description = ''
    error_cat_image = ''
    instance = Category.objects.get(pk=pk)
    if request.POST:
        form = AddCategoryForm(request.POST, request.FILES, instance=instance)
        if form.is_valid():
            form.save()
            return redirect(category_management)
        else:
            error_category_name = form['category_name'].errors
            error_slug = form['slug'].errors
            error_description = form['description'].errors
            error_cat_image = form['cat_image'].errors
    form = AddCategoryForm(instance=instance)
    context = {'form': form, 
               'error_category_name':error_category_name,
               'error_slug':error_slug,
               'error_description':error_description,
               'error_cat_image': error_cat_image,
               }
    return render(request, 'admin_panel/edit_category.html',context)


@user_passes_test(is_user_admin, login_url='admin_login')
def category_search(request):
    cat_obj=None
    if request.POST:
        search_item = request.POST.get('search_input')
        if search_item:
            cat_obj = Category.objects.order_by('id').filter(Q(category_name__icontains=search_item))
        context = {'cat_obj':cat_obj}
        return render(request, 'admin_panel/category_management.html',context)
    else:
        return redirect(category_management)


@never_cache
@user_passes_test(is_user_admin, login_url='admin_login')
def product_management(request):
    prod_obj = Product.objects.all().order_by('id')
    context = {'prod_obj' : prod_obj}
    return render(request, 'admin_panel/product_management.html',context)


@user_passes_test(is_user_admin, login_url='admin_login')
def unlist_product(request, pk):
    instance = Product.objects.get(pk=pk)
    instance.is_active=False
    instance.save()
    return redirect('product_management')


@user_passes_test(is_user_admin, login_url='admin_login')
def list_product(request, pk):
    instance = Product.objects.get(pk=pk)
    instance.is_active=True
    instance.save()
    return redirect('product_management')


@user_passes_test(is_user_admin, login_url='admin_login')
def add_product(request):
    error_product_name= ''
    error_slug= ''
    error_description= ''
    error_price= ''
    error_category= ''
    error_product_image= ''
    if request.POST:
        form = AddProductForm(request.POST)
        image_form = ProductImageForm(request.POST, request.FILES)
        if form.is_valid() and image_form.is_valid():
            product = form.save()
            for img in request.FILES.getlist('image'):
                image = ProductImages(image=img, product=product)
                image.save()
            return redirect(product_management)
        else:
            error_product_name = form['product_name'].errors
            error_slug = form['slug'].errors
            error_description = form['description'].errors
            error_price = form['price'].errors
            error_category = form['category'].errors
            error_product_image = image_form['image'].errors
    form = AddProductForm()
    image_form = ProductImageForm()
    context = {
        'form' : form,
        'image_form' : image_form,
        'error_product_name': error_product_name,
        'error_slug': error_slug,
        'error_description': error_description,
        'error_price': error_price,
        'error_category': error_category,
        'error_product_image': error_product_image,
        }
    return render(request, 'admin_panel/add_product.html',context)


@user_passes_test(is_user_admin, login_url='admin_login')
def edit_product(request, pk):
    product = get_object_or_404(Product, id=pk)
    error_product_name = ''
    error_slug = ''
    error_description = ''
    error_price = ''
    error_category = ''
    error_product_image = ''
    
    if request.method == 'POST':
        form = AddProductForm(request.POST, instance=product)
        image_form = ProductImageForm(request.POST, request.FILES)
        
        # If no image is provided, removing image field from form validation
        if not request.FILES:
            form.fields.pop('image', None)
            image_form.fields.pop('image', None)

        if form.is_valid() and image_form.is_valid():
            form.save()
            for img in request.FILES.getlist('image'):
                image = ProductImages(image=img, product=product)
                image.save()
            
            # For deleting the images
            if 'delete_images' in request.POST:
                images_to_delete = request.POST.getlist('delete_images')
                for image_id in images_to_delete:
                    image_to_delete = ProductImages.objects.get(id=image_id)
                    image_to_delete.delete()
            
            return redirect(product_management)
        else:
            error_product_name = form['product_name'].errors
            error_slug = form['slug'].errors
            error_description = form['description'].errors
            error_price = form['price'].errors
            error_category = form['category'].errors
            error_product_image = image_form.errors 
            
    else:
        form = AddProductForm(instance=product)
        image_form = ProductImageForm()
    
    context = {
        'form': form,
        'image_form': image_form,
        'error_product_name': error_product_name,
        'error_slug': error_slug,
        'error_description': error_description,
        'error_price': error_price,
        'error_category': error_category,
        'error_product_image': error_product_image,
        'product': product,
    }
    return render(request, 'admin_panel/edit_product.html', context)


@user_passes_test(is_user_admin, login_url='admin_login')
def product_search(request):
    prod_obj = None
    if request.POST:
        search_item = request.POST.get('search_input')
        if search_item:
            prod_obj = Product.objects.order_by('id').filter(Q(product_name__icontains=search_item))
        context = {'prod_obj':prod_obj}
        return render(request, 'admin_panel/product_management.html',context)
    else:
        return redirect(product_management)
    

@never_cache
@user_passes_test(is_user_admin, login_url='admin_login')
def variant_management(request):
    var_obj = Variation.objects.all().order_by('id')
    context = {'var_obj' : var_obj}
    return render(request, 'admin_panel/variant_management.html',context)



@user_passes_test(is_user_admin, login_url='admin_login')
def add_variant(request):
    form = AddVariantForm()
    if request.POST:
        form = AddVariantForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect(variant_management)
    context = {'form' : form}
    return render(request, 'admin_panel/add_variant.html',context)


@user_passes_test(is_user_admin, login_url='admin_login')
def edit_variant(request, pk):
    instance = Variation.objects.get(pk=pk)
    if request.POST:
        form = AddVariantForm(request.POST,instance=instance)
        if form.is_valid():
            form.save()
            return redirect(variant_management)
    form = AddVariantForm(instance=instance)
    context = {'form': form}
    return render(request, 'admin_panel/edit_variant.html',context)


@user_passes_test(is_user_admin, login_url='admin_login')
def variant_search(request):
    var_obj=None
    if request.POST:
        search_item = request.POST.get('search_input')
        if search_item:
            var_obj = Variation.objects.order_by('id').filter(Q(product__product_name__icontains=search_item))
        context = {'var_obj':var_obj}
        return render(request, 'admin_panel/variant_management.html',context)
    else:
        return redirect(variant_management)
    

@user_passes_test(is_user_admin, login_url='admin_login')
def coupon_management(request):
    coupon_obj = Coupon.objects.all().order_by('id')
    context = {'coupon_obj' : coupon_obj}
    return render(request, 'admin_panel/coupon_management.html',context)


@user_passes_test(is_user_admin, login_url='admin_login')
def add_coupon(request):
    form = AddCouponForm()
    if request.POST:
        form = AddCouponForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect(coupon_management)
    context = {'form' : form}
    return render(request, 'admin_panel/add_coupon.html',context)


@user_passes_test(is_user_admin, login_url='admin_login')
def unlist_coupon(request, pk):
    instance = Coupon.objects.get(pk=pk)
    instance.is_active=False
    instance.save()
    return redirect('coupon_management')


@user_passes_test(is_user_admin, login_url='admin_login')
def list_coupon(request, pk):
    instance = Coupon.objects.get(pk=pk)
    instance.is_active=True
    instance.save()
    return redirect('coupon_management')


@user_passes_test(is_user_admin, login_url='admin_login')
def edit_coupon(request, pk):
    instance = Coupon.objects.get(pk=pk)
    if request.POST:
        form = AddCouponForm(request.POST,instance=instance)
        if form.is_valid():
            form.save()
            return redirect(coupon_management)
    form = AddCouponForm(instance=instance)
    context = {'form': form}
    return render(request, 'admin_panel/edit_coupon.html',context)


@user_passes_test(is_user_admin, login_url='admin_login')
def order_management(request):
    order_obj = Order.objects.all().filter(is_ordered=True).order_by('-id')
    context = {'order_obj' : order_obj}
    return render(request, 'admin_panel/order_management.html',context)


@user_passes_test(is_user_admin, login_url='admin_login')
def order_search(request):
    order_obj=None
    if request.POST:
            search_item = request.POST.get('search_input')
            if search_item:
                order_obj = Order.objects.order_by('id').filter(Q(order_number__icontains=search_item)|Q(user__username__icontains=search_item)|Q(user__email__icontains=search_item))
            context = {'order_obj':order_obj}
            return render(request, 'admin_panel/order_management.html',context)
    else:
        return redirect(order_management)

@user_passes_test(is_user_admin, login_url='admin_login')
def order_details(request, order_id):
    order = Order.objects.get(id=order_id)
    order_subtotal = order.order_total - order.tax

    order_update_form = OrderUpdateForm(instance=order)
    payment_update_form = PaymentUpdateForm(instance=order)
    if request.POST:
        order_update_form = OrderUpdateForm(request.POST, instance=order)
        if order_update_form.is_valid():
            order_update_form.save()
            return redirect('order_management')


    context = {
        'order' : order,
        'order_subtotal' : order_subtotal,
        'order_update_form':order_update_form,
        'payment_update_form':payment_update_form,
        }
    return render(request, 'admin_panel/order_details.html',context)


@user_passes_test(is_user_admin, login_url='admin_login')
def cancel_order(request, order_id):
    order_obj = Order.objects.get(id=order_id)
    order_obj.status = 'Cancelled'
    order_obj.save()
    if order_obj.payment.status == 'Completed':
        wallet = Wallet.objects.get(user=request.user)
        wallet.balance += order_obj.order_total
        wallet.save()
        wallet_transaction = WalletTransaction.objects.create(
            transaction_id=str(uuid.uuid4().int)[:12],
            wallet=wallet,
            amount=order_obj.order_total,
            transaction_type='credit',
            order_reference=order_obj,
            updated_balance=wallet.balance,
        )
    
    order_items = order_obj.orderproduct_set.all()
    print(order_items)
    for item in order_items:
        product_variation = item.variation.all()

        #update the stock
        product = Product.objects.get(id=item.product_id)
        product_variations = product.variant.all()
        quantity = item.quantity
        for variation in product_variations:
            for i in product_variation:
                if i==variation:
                    variation.stock += quantity
                    variation.save()


    #order confirmation email
    subject = 'Order cancellation'
    message = f"""
    Hi {request.user.username},
    Your order has been cancelled!
    Order Number = {order_obj.order_number} 
    """
    send_mail(subject, message, "abdullamirshadcl@gmail.com", [request.user.email,], fail_silently=False)

    return redirect('order_management')

@user_passes_test(is_user_admin, login_url='admin_login')
def offer_management(request):
    offers = Offer.objects.all().order_by('-id')
    context = {
        'offers':offers,
    }
    return render(request, 'admin_panel/offer_management.html', context)

@user_passes_test(is_user_admin, login_url='admin_login')
def add_offer(request):
    form = AddOfferForm()
    if request.POST:
        form = AddOfferForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect(offer_management)
        
    context = {'form' : form}
    return render(request, 'admin_panel/add_offer.html', context)

@user_passes_test(is_user_admin, login_url='admin_login')
def edit_offer(request, offer_id):
    instance = Offer.objects.get(pk=offer_id)
    if request.POST:
        form = AddOfferForm(request.POST,instance=instance)
        if form.is_valid():
            form.save()
            return redirect(offer_management)
    form = AddOfferForm(instance=instance)
    context = {'form': form}
    return render(request, 'admin_panel/edit_offer.html', context)


@user_passes_test(is_user_admin, login_url='admin_login')
def deactivate_offer(request, pk):
    instance = Offer.objects.get(pk=pk)
    instance.is_active=False
    instance.is_sale=False
    instance.save()
    return redirect('offer_management')


@user_passes_test(is_user_admin, login_url='admin_login')
def activate_offer(request, pk):
    instance = Offer.objects.get(pk=pk)
    instance.is_active=True
    instance.is_sale=True
    instance.save()
    return redirect('offer_management')

@user_passes_test(is_user_admin, login_url='admin_login')
def sales_report(request):
    all_orders = Order.objects.all()
   
    start_date_str = '1900-01-01'
    end_date_str = '9999-12-31'
    if request.method == 'POST':
        start_date_str = request.POST.get('start', '1900-01-01')
        end_date_str = request.POST.get('end', '9999-12-31')
   

    start_date = datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else None
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d') if end_date_str else None

    if start_date is not None and end_date is not None:
        completed_order_items = OrderProduct.objects.filter(
            payment__status='Completed',
            order__created_at__gte=start_date,
            order__created_at__lte=end_date
        )
    else:
        completed_order_items = OrderProduct.objects.filter(
            payment__status='Completed',
        )

    sold_items = completed_order_items.values('product__product_name').annotate(total_sold=Sum('quantity'))

    total_sales = sum(order.order_total for order in all_orders)
    total_profit = int(sum(order_item.product.price * 0.3 for order_item in completed_order_items))

    most_sold_products = (
        completed_order_items.values('product__product_name')
        .annotate(total_sold=Sum('quantity'))
        .order_by('-total_sold')[:6]
    )
    most_sold_products_data = [
    {'product_name': item['product__product_name'], 'count': item['total_sold']}
    for item in most_sold_products
]


    product_profit_data = []

    for sold_stock_item in completed_order_items:
        product_price = sold_stock_item.product.price
        total_revenue = sold_stock_item.get_total
        total_revenue = int(total_revenue)
        profit = total_revenue - (total_revenue * 0.7)
        profit = int(profit)
        product_profit_data.append({
            'product_product_name': sold_stock_item.product.product_name,
            'total_sold': sold_stock_item.quantity,
            'profit': profit,
        })

    context = {
        'total_sales': total_sales,
        'most_sold_products': most_sold_products,
        'total_profit': total_profit,
        'sold_items': sold_items,
        'stock_items': product_profit_data,
        'start_date': start_date.strftime('%Y-%m-%d') if start_date else None,
        'end_date': end_date.strftime('%Y-%m-%d') if end_date else None,
    }
        
    return render(request, 'admin_panel/sales_report.html', context)


@user_passes_test(is_user_admin, login_url='admin_login')
def sales_report_pdf(request):
    
    all_orders = Order.objects.all()
    start_date_str = request.GET.get('start', '1900-01-01')
    end_date_str = request.GET.get('end', '9999-12-31')

    start_date = datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else None
    end_date = datetime.strptime(end_date_str, '%Y-%m-%d') if end_date_str else None
      
    completed_order_items = OrderProduct.objects.filter(
            payment__status='Completed',
            order__created_at__gte=start_date,
            order__created_at__lte=end_date
        )
    total_revenue = sum(order_item.get_total for order_item in completed_order_items)
    total_sales = sum(order_item.quantity for order_item in completed_order_items)
    total_profit = sum(order_item.product.price * 0.3 for order_item in completed_order_items)
    
    all_orders = Order.objects.filter().distinct()
    
    sold_items = [{
        'product_name': item.product.product_name,
        'total_sold': item.quantity,
        'profit': int(item.get_total - (item.get_total * 0.7)),
    } for item in completed_order_items]
    sold_items = sorted(sold_items, key=lambda x: x['total_sold'], reverse=True)

    context = {
        'start_date':start_date,
        'end_date':end_date,
        'total_sales': total_sales,
        'total_revenue':total_revenue,
        'total_profit':total_profit,
        'sold_items':sold_items,
        'all_orders':all_orders,
        'delivered_order_items':completed_order_items,
        
    }

    pdf = render_to_pdf('admin_panel/sales_report_pdf.html', context)

    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        filename = "Sales Report.pdf"
        content = f'attachment; filename="{filename}"'
        response['Content-Disposition'] = content
        return response

    return HttpResponse('Failed to generate PDF')

@user_passes_test(is_user_admin, login_url='admin_login')
def sales_report_excel(request):
    try:
        start_date_str = request.GET.get('start', '1900-01-01')
        end_date_str = request.GET.get('end', '9999-12-31')
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d') if start_date_str else None
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d') if end_date_str else None
        completed_order_items = OrderProduct.objects.filter(
            payment__status='Completed',
            order__created_at__gte=start_date,
            order__created_at__lte=end_date
        )

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet['A1'] = 'Total Revenue'
        worksheet['B1'] = sum(order_item.get_total for order_item in completed_order_items)
        worksheet['A2'] = 'Total Sales'
        worksheet['B2'] = sum(order_item.quantity for order_item in completed_order_items)
        worksheet['A3'] = 'Total Profit'
        worksheet['B3'] = sum(order_item.product.price * 0.3 for order_item in completed_order_items)

        row_num = 5
        headers = ['Order ID', 'Date Ordered', 'Total Amount', 'Products']
        for col_num, header in enumerate(headers, 1):
            worksheet.cell(row=row_num, column=col_num, value=header)

        row_num += 1
        date_style = NamedStyle(name='date_style', number_format='YYYY-MM-DD')
        for order_item in completed_order_items:
            worksheet.cell(row=row_num, column=1, value=order_item.order.id)

            date_ordered_naive = order_item.order.created_at.astimezone(timezone.utc).replace(tzinfo=None)
            worksheet.cell(row=row_num, column=2, value=date_ordered_naive)
            worksheet.cell(row=row_num, column=2).style = date_style

            worksheet.cell(row=row_num, column=3, value=order_item.order.order_total)

            # Create a list of formatted strings for products
            products_list = [f"{order_item.product.product_name} ({order_item.quantity} units) - ${order_item.quantity * order_item.product_price}"]
            
            # Join the list with newline characters
            products_str = '\n'.join(products_list)
            
            # Write the formatted string to the worksheet cell
            worksheet.cell(row=row_num, column=4, value=products_str)

            # Increment row number for the next iteration
            row_num += 1  # Assuming you have this increment in your code


        for i in range(1, 4):
            cell = worksheet.cell(row=i, column=2)
            cell.font = openpyxl.styles.Font(bold=True)

        header_font = openpyxl.styles.Font(bold=True, color='FFFFFF')
        header_alignment = openpyxl.styles.Alignment(horizontal='center')
        for col_num, header in enumerate(headers, 1):
            cell = worksheet.cell(row=row_num - len(completed_order_items) - 1, column=col_num)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.fill = openpyxl.styles.PatternFill(start_color='007bff', end_color='007bff', fill_type='solid')

        data_alignment = openpyxl.styles.Alignment(wrap_text=True)
        for row_num in range(row_num - len(completed_order_items), row_num):
            for col_num in range(1, 5):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.alignment = data_alignment
                cell.border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
                
        for col_num, header in enumerate(headers, 1):
            max_length = len(header)
            for row_num in range(2, row_num):
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                try:
                    if len(str(cell_value)) > max_length:
                         max_length = len(cell_value)
                except:
                        pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[get_column_letter(col_num)].width = adjusted_width

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=Sales_Report.xlsx'
        workbook.save(response)

        return response

    except Exception as e:
        print(f"Error generating Excel file: {e}")
        return HttpResponse("Failed to generate Excel file", status=500)
    
@user_passes_test(is_user_admin, login_url='admin_login')
def get_sales_data(request, period):
    if period == 'week':
        
        start_date = tz.now().date() - tz.timedelta(days=6)
        order_items = OrderProduct.objects.filter(order__created_at__gte=start_date)
        data = (
            order_items.annotate(day=TruncDay('order__created_at'))
            .values('day')
            .annotate(total=Sum(F('quantity') * F('product__price')))
            .order_by('day')
        )
        labels = [item['day'].strftime('%A') for item in data]
    elif period == 'month':
        start_date = tz.now().date() - tz.timedelta(days=30)
        order_items = OrderProduct.objects.filter(order__created_at__gte=start_date)
        data = (
        order_items.annotate(day=TruncDay('order__created_at'))
        .values('day')
        .annotate(total=Sum(F('quantity') * F('product__price')))
        .order_by('day')
    )
        labels = [item['day'].strftime('%Y-%m-%d') for item in data]
    elif period == 'year':
        start_date = tz.now().date() - tz.timedelta(days=365)
        order_items = OrderProduct.objects.filter(order__created_at__gte=start_date)
        data = (
            order_items.annotate(month=TruncMonth('order__created_at'))
            .values('month')
            .annotate(total=Sum(F('quantity') * F('product__price')))
            .order_by('month')
        )
        labels = [f"{item['month'].strftime('%B')}" for item in data]
    else:
        return JsonResponse({'error': 'Invalid period'})
    
    sales_data = [item['total'] for item in data]
    return JsonResponse({'labels': labels, 'data': sales_data})